'''
References:

Setting up the environment:
Note: Playwright requires python3.7 and above
      If a virtualenv had been created using an older version, just delete the virtualenv folder
      and create a new one.
virtualenv -p python3.8 rpa_playwright
source rpa_playwright/bin/activate
pip3 install openpyxl
pip3 install playwright
playwright install
pip3 freeze > requirements.txt
'''

from openpyxl import load_workbook
from playwright.sync_api import sync_playwright

# load the workbook and specify the sheet
wb = load_workbook(filename='challenge.xlsx')
sheet = wb['Sheet1']

# get workbook range information
max_rows = sheet.max_row
max_cols = sheet.max_column
print("Max rows: " + str(max_rows))
print("Max cols: " + str(max_cols))

# input the values to rpachallenge website using playwright
with sync_playwright() as p:
    browser = p.chromium.launch(headless=False)
    page = browser.new_page()
    page.goto('http://www.rpachallenge.com')
    page.click('button.waves-effect')
    for r in range(2, max_rows + 1):
        # first name
        page.type('//*[@ng-reflect-name="labelFirstName"]', sheet.cell(row=r, column=1).value)
        # last name
        page.type('//*[@ng-reflect-name="labelLastName"]', sheet.cell(row=r, column=2).value)
        # company name
        page.type('//*[@ng-reflect-name="labelCompanyName"]', sheet.cell(row=r, column=3).value)
        # role in company
        page.type('//*[@ng-reflect-name="labelRole"]', sheet.cell(row=r, column=4).value)
        # address
        page.type('//*[@ng-reflect-name="labelAddress"]', sheet.cell(row=r, column=5).value)
        # email
        page.type('//*[@ng-reflect-name="labelEmail"]', sheet.cell(row=r, column=6).value)
        # phone number
        page.type('//*[@ng-reflect-name="labelPhone"]', str(sheet.cell(row=r, column=7).value))
        # submit
        page.click('input.btn')
    page.screenshot(path='rpa-challenge-result.png')
    browser.close()
